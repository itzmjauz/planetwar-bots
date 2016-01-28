from openpyxl.styles import Font, Color, colors
from openpyxl import Workbook, load_workbook
import sys, time, string

"""
Example:
Running BerendHillBot.py vs AdaptiveBot.py on maps/3planets/map1.txt
Game[1]: Player 2 Wins!
Game took Game state turn 9 turns, had        0 timeouts

('maps/3planets/map1.txt', 'BerendHillBot.py', 'AdaptiveBot.py', -1, '9')
"""

def AddToSheet(filename, sheets, coords, columns):
    file = open(filename).read().split("\n")

    results = sorted([extract(i) for i in range(0,len(file)-2,3)], key=lambda x: x[3])
    maps = sorted([(games[i][0], games[i][-2]+games[i+1][-2]+games[i+2][-2]+games[i+3][-2]+games[i+4][-2]) for i in range(0, len(games), 5)], key=lambda x: x[1], reverse=True)
    print("Processing file {0}: ".format(filename))
    print("Processing {0} Results...".format(len(games)))
    t = time.time()
    wb = load_workbook("Template.xlsx")
    last_wb = "Template.xlsx"

    for data in results:
        print(data)
        if (last_wb!=sheets[data[2]]):
            wb.save(last_wb)
            last_wb = sheets[data[2]]
            wb = load_workbook(sheets[data[2]])
            ws = wb.active
        column = columns[filename.split("_")[2]]+int(filename.split(".")[0][-1])
        cell = string.ascii_uppercase[column]+str(coords[data[0]])
        ws[cell] = data[-1]
        if (data[-2]==-1):
            ws[cell].font= Font(color=colors.RED)
        if (data[-2]==1):
            ws[cell].font= Font(color=colors.GREEN)
    print("Processed in {0} Seconds".format(str(time.time()-t)[0:5]))

def extract(index):
    lines = [file[index], file[index+1], file[index+2]]
    player = lines[0].split(" vs ")[0].split(" ")[-1]
    enemy = lines[0].split(" vs ")[1].split(" ")[0]
    map = lines[0].split(" vs ")[1].split(" ")[-1]

    if "Draw!" in lines[1]:
        state = 0
    else:
        if "Player 1" in lines[1]:
            state = 1
        else:
            state = -1
    turns = lines[2].split(" ")[5]
    return (map, player, enemy, state, turns)

file = open(sys.argv[1]).read().split("\n")
print(len(file))

rows = dict(zip(['maps/3planets/map1.txt','maps/3planets/map2.txt','maps/3planets/map3.txt','maps/4planets/map1.txt','maps/4planets/map2.txt','maps/4planets/map3.txt','maps/5planets/map1.txt','maps/5planets/map2.txt','maps/5planets/map3.txt','maps/6planets/map1.txt','maps/6planets/map2.txt','maps/6planets/map3.txt','maps/7planets/map1.txt','maps/7planets/map2.txt','maps/7planets/map3.txt','maps/8planets/map1.txt','maps/8planets/map2.txt','maps/8planets/map3.txt'], range(4,22)))

columns = {"hill": 1, "minmax": 6, "alphabeta":11}

sheets = {'BullyBot.py':"Bully.xlsx", 'RandomBot.py':"Random.xlsx", 'EmptyBot.py':"Empty.xlsx", 'LookaheadBot.py':"Lookahead.xlsx", 'AdaptiveBot.py':"Adaptive.xlsx"}

for i in argv[1:]:
    AddToSheet(i, sheets, rows, columns)
